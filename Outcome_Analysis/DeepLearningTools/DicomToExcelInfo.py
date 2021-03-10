#Script written by Raymond Mumme, RPMumme@mdanderson.org
#Script edited by Brian Anderson, bmanderson@mdanderson.org
#Goes through each patient directory and finds their CT file sets, then collects data from the dicoms.
#Writes out a new line on the spreadsheet for every new set.
#Highlights duplicate patient names
#
#Change the path variable to the directory containing all the patient directories
#Change the sheet_path variable to the location/file name of the spreadsheet
#
#Note that for the packages, openpyxl comes from pandas

import os
import glob
import re
import pydicom as dicom

import pandas
from openpyxl import Workbook, load_workbook
from openpyxl import styles
from openpyxl import formatting
from datetime import date
import time
import logging


def return_workbook(sheet_path):
    if os.path.exists(sheet_path):
        wb = load_workbook(sheet_path)
    else:
        # Workbook is created
        wb = Workbook()
        # add_sheet is used to create sheet
    sheet1 = wb.active
    sheet1.title = 'Sheet1'

    # adding column headers
    sheet1.cell(1, 1).value = 'Patient'
    sheet1.cell(1, 2).value = 'Date'
    sheet1.cell(1, 3).value = 'Age'
    sheet1.cell(1, 4).value = 'Gender'
    sheet1.cell(1, 5).value = 'Width'
    sheet1.cell(1, 6).value = 'Height'
    sheet1.cell(1, 7).value = 'Rescale'
    sheet1.cell(1, 8).value = 'Data Diam'
    sheet1.cell(1, 9).value = 'Recon Diam'
    sheet1.cell(1, 10).value = 'Filter'
    sheet1.cell(1, 11).value = 'Kernel'
    sheet1.cell(1, 12).value = 'Options'
    sheet1.cell(1, 13).value = 'Manufacturer'
    sheet1.cell(1, 14).value = 'Model'
    sheet1.cell(1, 15).value = 'Photometric'
    sheet1.cell(1, 16).value = 'Spacing'
    sheet1.cell(1, 17).value = 'Thickness'
    sheet1.cell(1, 18).value = 'X Pixel Sp'
    sheet1.cell(1, 19).value = 'Y Pixel Sp'
    sheet1.cell(1, 20).value = 'kVp'
    sheet1.cell(1, 21).value = 'Current'
    sheet1.cell(1, 22).value = 'Image Difs'
    sheet1.cell(1, 23).value = 'Avg Dif'
    sheet1.cell(1, 24).value = 'Slices'
    sheet1.cell(1, 25).value = 'Pat Position'
    sheet1.cell(1, 26).value = 'Passed?'
    return wb


def add_files_to_sheet(new_files, sheet1, n=None):
    if n is None:
        n = sheet1.max_row + 1
    dcm_info = dicom.read_file(new_files[0], force=True)
    if dcm_info.Modality == 'CT':
        sheet1.cell(n, 1).value = dcm_info.PatientID
        sheet1.cell(n, 26).value = 'False'
        pixels_x = None
        pixels_y = None
        spacing = None
        thickness = None
        age = None
        studydate = None
        bdate = None
        gender = None
        patpos = None
        filtertype = None
        kernel = None
        options = None
        manufact = None
        model = None
        photomet = None
        kvp = None
        current = None
        width = None
        height = None
        rescale = None
        datadiam = None
        recondiam = None
        num_slices = None

        study_bool = False
        birth_bool = False

        z_ct_locs = []

        pixels_x = dcm_info.PixelSpacing[0]
        pixels_y = dcm_info.PixelSpacing[1]
        if hasattr(dcm_info, 'SliceThickness'):
            thickness = dcm_info.SliceThickness
        if hasattr(dcm_info, 'StudyDate'):
            studydate = dcm_info.StudyDate
            studydate = studydate[4:6] + '/' + studydate[6:8] + '/' + studydate[:4]
            if len(studydate) > 5:
                study_bool = True

        if hasattr(dcm_info, 'PatientBirthDate'):
            bdate = dcm_info.PatientBirthDate
            bdate = bdate[4:6] + '/' + bdate[6:8] + '/' + bdate[:4]
            if len(bdate) > 5 and (bdate != studydate):
                birth_bool = True

        if hasattr(dcm_info, 'PatientAge'):
            age = dcm_info.PatientAge
            age = ''.join(age.split())
            if age and ('-' not in age):
                age = int(''.join(filter(lambda x: x.isdigit(), age)))
                if age < 1 or age > 130:
                    if birth_bool and study_bool:
                        age = calculate_age(bdate, studydate)
                        if age > 130:
                            age = None
                    else:
                        age = None
            else:
                age = None
        else:
            if birth_bool and study_bool:
                age = calculate_age(bdate, studydate)
                if age > 130:
                    age = None
            else:
                age = None

        if hasattr(dcm_info, 'PatientSex'):
            if birth_bool:
                gender = dcm_info.PatientSex

        width = dcm_info.Rows
        height = dcm_info.Columns
        rescale = dcm_info.RescaleIntercept
        if hasattr(dcm_info, 'DataCollectionDiameter'):
            datadiam = dcm_info.DataCollectionDiameter
        if hasattr(dcm_info, 'ReconstructionDiameter'):
            recondiam = dcm_info.ReconstructionDiameter
        if hasattr(dcm_info, 'FilterType'):
            filtertype = dcm_info.FilterType
        if hasattr(dcm_info, 'ConvolutionKernel'):
            kernel = dcm_info.ConvolutionKernel
        if hasattr(dcm_info, 'ScanOptions'):
            options = dcm_info.ScanOptions
        if hasattr(dcm_info, 'Manufacturer'):
            manufact = dcm_info.Manufacturer
        if hasattr(dcm_info, 'ManufacturerModelName'):
            model = dcm_info.ManufacturerModelName
        if hasattr(dcm_info, 'PhotometricInterpretation'):
            photomet = dcm_info.PhotometricInterpretation
        if hasattr(dcm_info, 'KVP'):
            kvp = dcm_info.KVP
        if hasattr(dcm_info, 'XRayTubeCurrent'):
            current = dcm_info.XRayTubeCurrent
        if hasattr(dcm_info, 'PatientPosition'):
            patpos = dcm_info.PatientPosition
        if hasattr(dcm_info, 'SpacingBetweenSlices'):
            spacing = dcm_info.SpacingBetweenSlices
            if spacing is not None:
                spacing = abs(int(spacing))

        for file in new_files:
            dcm_info = dicom.read_file(file, force=True)
            if dcm_info.Modality == 'CT':
                z_ct_locs.append(float(dcm_info.ImagePositionPatient[2]))

        # sort z locations for correct differences
        z_ct_locs.sort()
        difs = []
        for m in range(0, len(z_ct_locs) - 1):
            difs.append(abs(z_ct_locs[m + 1] - z_ct_locs[m]))
        avg_dif = round(sum(difs) / len(difs), 2)
        num_slices = len(z_ct_locs)
        sheet1.cell(n, 2).value = studydate
        sheet1.cell(n, 3).value = age
        sheet1.cell(n, 4).value = gender
        sheet1.cell(n, 5).value = width
        sheet1.cell(n, 6).value = height
        sheet1.cell(n, 7).value = rescale
        sheet1.cell(n, 8).value = datadiam
        sheet1.cell(n, 9).value = recondiam
        sheet1.cell(n, 10).value = filtertype
        sheet1.cell(n, 11).value = kernel
        sheet1.cell(n, 12).value = options
        sheet1.cell(n, 13).value = manufact
        sheet1.cell(n, 14).value = model
        sheet1.cell(n, 15).value = photomet
        sheet1.cell(n, 16).value = spacing
        sheet1.cell(n, 17).value = thickness
        sheet1.cell(n, 18).value = pixels_x
        sheet1.cell(n, 19).value = pixels_y
        sheet1.cell(n, 20).value = kvp
        sheet1.cell(n, 21).value = current
        sheet1.cell(n, 22).value = str(difs)
        sheet1.cell(n, 23).value = avg_dif
        sheet1.cell(n, 24).value = num_slices
        sheet1.cell(n, 25).value = patpos
        sheet1.cell(n, 26).value = 'True'


def calculate_age(birthdate,studydate):
    studydate_s = studydate.split('/')
    studydate_a = []
    birthdate_s = birthdate.split('/')
    birthdate_a = []

    for num in studydate_s:
        studydate_a.append(int(num))
    for num in birthdate_s:
        birthdate_a.append(int(num))


    study = date(studydate_a[2],studydate_a[0],studydate_a[1])
    birth = date(birthdate_a[2],birthdate_a[0],birthdate_a[1])

    return study.year - birth.year - ((study.month, study.day) < (birth.month, birth.day))


def add_patients_from_path(path, wb):
    sheet1 = wb.active
    for i, (root, dirs, files) in enumerate(os.walk(path)):
        # gather the information from each CT file
        new_files = glob.glob(os.path.join(root, '*.dcm'))
        if len(new_files) > 10:
            print(root)
            finished = False
            retry = 0
            n = sheet1.max_row + 1
            while not finished:
                try:
                    add_files_to_sheet(new_files=new_files, sheet1=sheet1, n=n)
                    finished = True
                except BaseException as e:
                    if retry > 0 or (not isinstance(e,Exception)):
                        logging.exception(e)
                        finished = True
                    else:
                        logging.exception(e)
                        print('Iteration: i = ', i)
                        print('Trying again...')
                        retry += 1
                        time.sleep(3)


def main():
    yellow_color = 'FFFF00'
    yellow_fill = styles.PatternFill(start_color=yellow_color, end_color=yellow_color, fill_type='solid')
    dxf = styles.differential.DifferentialStyle(fill=yellow_fill)
    duplicate_rule = formatting.Rule(type="duplicateValues", dxf=dxf, stopIfTrue=None)

    sheet_path = r'K:\Morfeus\BMAnderson\CNN\Data\Data_Liver\DicomInfo.xlsx'
    paths = [r'K:\Morfeus\BMAnderson\CNN\Data\Data_Liver\Miccai_Challenge',
             r'K:\Morfeus\BMAnderson\CNN\Data\Data_Liver\Data'
             ]
    wb = return_workbook(sheet_path=sheet_path)
    for path in paths:
        add_patients_from_path(path=path, wb=wb)
    sheet1 = wb.active
    n = sheet1.max_row
    sheet1.conditional_formatting.add('A2:A'+str(n), duplicate_rule)
    wb.save(sheet_path)
    print('Complete spreadsheet saved.')


if __name__ == '__main__':
    main()
